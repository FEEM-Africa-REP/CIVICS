# -*- coding: utf-8 -*-
"""
Created on Mon Aug 24 12:12:55 2020

@author: Mohammad Amin Tahavori
"""

import REP_CVX as cvx

a=cvx.C_SUT(path=r'Database\Kenya_2014_SAM.xlsx',unit='M KSH') 
#%%
a.shock_calc(path=r'sensitivity\a1\q1.xlsx',Z=True)  
#%%
a.plot_dx()
#%%

#%%
import openpyxl
xl_file = r'shading_trees.xlsx'
myworkbook=openpyxl.load_workbook(xl_file)
worksheet= myworkbook.get_sheet_by_name('main')
rows = worksheet.max_row
cols = worksheet.max_column
sens_col = 5
val_col  = 3
par_col  = 2
min_col  = 6
max_col  = 7
stp_col  = 8
sensitivity_info = {}

counter = 0
for row in range(rows):
    if worksheet.cell(row=row+1, column=5).value == 'Yes':
        sensitivity_info['{}'.format(counter)] = {'parameter':worksheet.cell(row=row+1, column=par_col).value,
                                                  'minimum':worksheet.cell(row=row+1, column=min_col).value,
                                                  'maximum':worksheet.cell(row=row+1, column=max_col).value,
                                                  'step':worksheet.cell(row=row+1, column=stp_col).value,
                                                  'row':row+1}
        counter+=1

# myworkbook.save(xl_file)
#%%
import os
import shutil
directs=[]
for i in range (counter):
    dir = os.path.join('sens_{}'.format(sensitivity_info[str(i)]['parameter']))
    if not os.path.exists(dir):
        os.mkdir(dir)
    else:
        shutil.rmtree(dir)
        os.mkdir(dir)

        
    directs.append(dir)
    s_min = sensitivity_info[str(i)]['minimum']
    s_max = sensitivity_info[str(i)]['maximum']
    step  = sensitivity_info[str(i)]['step']
    row   = sensitivity_info[str(i)]['row']
    while s_min <= s_max:
        myworkbook=openpyxl.load_workbook(xl_file)
        worksheet= myworkbook.get_sheet_by_name('main')
        worksheet.cell(row=row, column=val_col).value = s_min
        name = directs[i] + '\{}.xlsx'.format(s_min)
        myworkbook.save(name)
        s_min+=step
        
        
        
        
    
#%%
   

def dict_maker(Z=None,X=None,VA=None,p=None,Y=None,va=None,z=None,s=None,
               Z_agg=None,X_agg=None,VA_agg=None,Y_agg=None,S_agg=None):
    
    inputs =  [Z,X,VA,p,Y,va,z,s,Z_agg,X_agg,VA_agg,Y_agg,S_agg]
    outputs = ['Z','X','VA','p','Y','va','z','s','Z_agg','X_agg','VA_agg','Y_agg','S_agg']
    
    dictionary = {}
    
    for i  in range(len(inputs)):
        if inputs[i] is not None:
            dictionary[outputs[i]]=inputs[i]
               
    return dictionary

 #%%   
import pandas as pd
h = pd.DataFrame(0,index=['a'],columns=['a'])
a = dict_maker(Z=h) 
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    