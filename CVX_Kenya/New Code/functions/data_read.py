# -*- coding: utf-8 -*-
"""
Created on Mon Aug 24 11:28:05 2020

@author: Mohammad Amin Tahavori
"""
def database(path):
    import pandas as pd
    
    SUT = pd.read_excel(path, index_col=[0,1,2,3,4], header = [0,1,2,3,4])
    
        # importing use (U), supply (V), supply-use together (Z) and satellite accounts (S)
    U = SUT.loc['Commodities','Activities']
    V = SUT.loc['Activities','Commodities']
    Z = SUT.loc[['Commodities','Activities'], ['Commodities','Activities']]
    S = SUT.loc['Satellite Accounts',['Commodities','Activities']]
        
        
        # computing total final demand (Y) by importing households (HH), investment (IN), government (GO) and export (EX) 
    HH  = SUT.loc[['Commodities','Activities'], 'Households']
    IN  = SUT.loc[['Commodities','Activities'], 'Savings-Investment']
    GO  = SUT.loc[['Commodities','Activities'], 'Government']
    EX  = SUT.loc[['Commodities','Activities'], 'Rest of the World']
    Y_M = SUT.loc[['Commodities','Activities'], 'Margins']
        
    Y = pd.DataFrame(HH.sum(axis=1) + IN.sum(axis=1) + GO.sum(axis=1) + EX.sum(axis=1) + Y_M.sum(axis=1), index=HH.index, columns=['Total final demand'])
        
        # computing total value added (VA) by importing factors of production (F), taxes (T), import (IM) and margins as factor (F_M)
    F   = SUT.loc['Factors', ['Commodities', 'Activities']]
    T   = SUT.loc['Government', ['Commodities','Activities']]
    IM  = SUT.loc['Rest of the World', ['Commodities','Activities']]
    F_M = SUT.loc['Margins', ['Commodities','Activities']]
        
    VA  = F.append(T.append(IM.append(F_M)))
        
        # computing total production vector (X)
    X = pd.DataFrame(Y.sum(axis=1) + Z.sum(axis=1), index=Z.index, columns=['Total Production'])
        
    return SUT,U,V,Z,S,Y,VA,X

def sens_info(path):
    
    import openpyxl
    import os
    import shutil


    myworkbook=openpyxl.load_workbook(path)
    worksheet= myworkbook.get_sheet_by_name('main')
    rows = worksheet.max_row
    
    sens_col = 5
    val_col  = 3
    par_col  = 2
    min_col  = 6
    max_col  = 7
    stp_col  = 8
    
    sensitivity_info = {}
    counter = 0
    for row in range(rows):
        if worksheet.cell(row=row+1, column=sens_col).value == 'Yes':

            sensitivity_info['{}'.format(counter)] = {'parameter':worksheet.cell(row=row+1, column=par_col).value,
                                                                 'minimum':worksheet.cell(row=row+1, column=min_col).value,
                                                                 'maximum':worksheet.cell(row=row+1, column=max_col).value,
                                                                 'step':worksheet.cell(row=row+1, column=stp_col).value,
                                                                 'row':row+1}
            counter+=1                                                    
    
    print ('{} sensitivities are found'.format(counter))
    
    directs=[]
    for i in range (counter):
        dir = os.path.join('sens_{}'.format(sensitivity_info[str(i)]['parameter']))
        if not os.path.exists(dir):
            os.mkdir(dir)
        else:
            shutil.rmtree(dir)
            os.mkdir(dir)
        directs.append(dir)
        s_min = sensitivity_info[str(i)]['minimum']
        s_max = sensitivity_info[str(i)]['maximum']
        step  = sensitivity_info[str(i)]['step']
        row   = sensitivity_info[str(i)]['row']
        
        while s_min <= s_max:
            myworkbook=openpyxl.load_workbook(path)
            worksheet= myworkbook.get_sheet_by_name('main')
            worksheet.cell(row=row, column=val_col).value = s_min
            name = directs[i] + '\case_{}.xlsx'.format(s_min)
            myworkbook.save(name)
            s_min+=step    
    
    return directs,sensitivity_info
        